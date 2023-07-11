import yfinance as yf
from openpyxl import load_workbook
import datetime as dt
from datetime import date
from datetime import timedelta

#Calculates return for the next day given hedging pressure of the previous day.
def get_sheetnames_xlsx(filepath): 
    wb = load_workbook(filepath, read_only=True, keep_links=False)
    return wb.sheetnames

sheet_name = get_sheetnames_xlsx('calls_output.xlsx')
sheet_name = sheet_name[0].split()

expiration_date = sheet_name[1]

ticker_text = sheet_name[0]
ticker = yf.Ticker(ticker_text)

stock_data = ticker.history()

wb = load_workbook("C:/Users/Administrator/Desktop/Mercy/vscode/SmartMoneyAlgo/hpReturnData.xlsx")
sheet = wb.active 
last_row = len(sheet['A'])
last_row_date = sheet["A" + str(last_row)]

stock_closes = stock_data.iloc[:, 3].values.tolist()
stock_at_30 = sheet["E" + str(last_row)].value
current_p = stock_closes[-1]
current_return =  ((current_p - stock_at_30) / stock_at_30) *100

today = date.today()

sheet["E" + str(last_row)] = round(current_return, 5)

wb.save("C:/Users/Administrator/Desktop/Mercy/vscode/SmartMoneyAlgo/hpReturnData.xlsx")
print("done")
