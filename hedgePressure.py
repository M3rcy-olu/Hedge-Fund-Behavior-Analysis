import pandas as pd
import yfinance as yf
from openpyxl import load_workbook
import datetime as dt
from datetime import date

with open('C:/Users/Mercy/Documents/GitHub/SmartMoneyAlgo/calcGreeks_final.py') as f: 
    exec(f.read())

#Dictionary to easily access the columns/indecies for each component (ex: strike price)
header_dict = {
    'contractSymbol': 0,
    'lastTradeDate' : 1,
    'strike': 2,
    'lastPrice': 3,
    'bid' : 4,
    'ask' : 5,
    'change' : 6,
    'percentChange' : 7,
    'volume' : 8,
    'openInterest' : 9, 
    'impliedVolatility' : 10,
    'inTheMoney' : 11, 
    'contractSize' : 12,
    'currency' : 13,
    'delta' : 14, 
    'gamma' : 15, 
    'vega' : 16, 
    'rho' : 17, 
    'theta' : 18

}

#Defines functions
def get_sheetnames_xlsx(filepath): 
    wb = load_workbook(filepath, read_only=True, keep_links=False)
    return wb.sheetnames

def calc_dol_gamma(option, option_type):
    gamma = option[header_dict['gamma']]
    open_interest = option[header_dict['openInterest']]
    if option_type == 'p': 
        dollar_gamma =  -1 * gamma * open_interest * 100 * underlying_p
    else: 
        dollar_gamma = gamma * open_interest * 100 * underlying_p
    return dollar_gamma

def calc_gamma_imb(ticker, call_dollar_gammas, put_dollar_gammas): 
    vol_history = ticker.history(period= '1mo')

    total_dol_vol = 0
    max_counter = len(vol_history)
    counter = 0 
    while (counter < max_counter): 
        vol_history[vol_history.columns[4]] = vol_history[vol_history.columns[4]].astype(int)
        vol = vol_history.iloc[counter, 4]
        vol_history[vol_history.columns[3]] = vol_history[vol_history.columns[3]].astype(int)
        price = vol_history.iloc[counter, 3]
        total_dol_vol += manage_const * vol * price
        counter += 1

    avg_dol_vol = total_dol_vol / (max_counter)


    total_call_g = 0 
    total_put_g = 0

    for g in call_dollar_gammas:
        total_call_g += g

    for m in put_dollar_gammas: 
        total_put_g += m 

    gamma_imb = (total_call_g + total_put_g) * (underlying_p/100) * (1/avg_dol_vol) * manage_const
    return gamma_imb

def calc_hedge_pressure(gamma_imb): 
    stock_closes = stock_data.iloc[:, 3].values.tolist()
    stock_opens = stock_data.iloc[:, 0].values.tolist()
    prev_close_p = stock_closes[-2]
    current_p = stock_closes[-1]
    current_return =  (current_p - prev_close_p) / prev_close_p
    hedge_pressure = 100 * gamma_imb * current_return
    return hedge_pressure, current_return

#Overall gathers data from excel sheets
#Initializes variables for stock name and experiation date. 
#Converts Excel to dataframe 
sheet_name = get_sheetnames_xlsx('calls_output.xlsx')
sheet_name = sheet_name[0].split()

expiration_date = sheet_name[1]

ticker_text = sheet_name[0]
ticker = yf.Ticker(ticker_text)

stock_data = ticker.history()
stock_price = stock_data.iloc[:, 3].values.tolist()
underlying_p = stock_price[-1]
manage_const = 1e-5

call_chain = pd.read_excel('calls_output.xlsx')
call_chain = call_chain.drop(call_chain.columns[0], axis = 1)
put_chain = pd.read_excel('puts_output.xlsx')
put_chain = put_chain.drop(put_chain.columns[0], axis = 1)

#Execution function to calculate dollar gammas 
#Dollar gamma = gamma * open interest * 100 * latest underlying price --> for an option on the experiation date from the excel sheet for each strike price
call_dollar_gammas = []
put_dollar_gammas = []

num_options = len(call_chain)
counter = 0
while (counter < num_options):
    call_option = call_chain.loc[counter].values.flatten().tolist()
    call_dol_gamma = calc_dol_gamma(call_option, 'c')
    call_dollar_gammas.append(call_dol_gamma)
    counter += 1

num_options = len(put_chain)
counter = 0
while (counter < num_options):
    put_option = put_chain.loc[counter].values.flatten()
    put_dol_gamma = calc_dol_gamma(put_option, 'p')
    put_dollar_gammas.append(put_dol_gamma)
    counter+=1

gamma_imb = calc_gamma_imb(ticker, call_dollar_gammas, put_dollar_gammas)
hedge_pressure, r_open_last30 = calc_hedge_pressure(gamma_imb)

#updating calculated values to data excel
today = date.today()
today_str = today.strftime("%m/%d/%y")

final_data = [today_str, round(gamma_imb, 5), round(hedge_pressure, 5), r_open_last30, underlying_p]

wb = load_workbook("C:/Users/Mercy/Documents/GitHub/SmartMoneyAlgo/hpReturnData.xlsx")
sheet = wb.active 
last_row = len(sheet['A'])
last_row_date = sheet["A" + str(last_row)].value

if last_row_date == final_data[0]: 
    sheet["B" + str(last_row)] = final_data[1]
    sheet["C" + str(last_row)] = final_data[2]
    sheet["D" + str(last_row)] = final_data[3]
    sheet["E" + str(last_row)] = final_data[4]
else: 
    sheet["A" + str(last_row+1)] = final_data[0]
    sheet["B" + str(last_row+1)] = final_data[1]
    sheet["C" + str(last_row+1)] = final_data[2]
    sheet["D" + str(last_row+1)] = final_data[3]
    sheet["E" + str(last_row+1)] = final_data[4]
wb.save("C:/Users/Mercy/Documents/GitHub/SmartMoneyAlgo/hpReturnData.xlsx")

print("part 2 done")
#Extra print and write to excells that I needed when testing my code. 
#These are not needed in the actual functionality of the code.
'''print(expiration_date)
print("dollar gamma imbalance is per 1% move in the underlying stock price")
print("Dollar Gamma Imbalance: " + str(round(gamma_imb, 5)))
print("Hedge Pressure: " + str(round(hedge_pressure, 5)))
print("Current Return: " + str(round(r_open_last30 * 100, 5)) + "%")

Combines calculated dollar gammas with Dataframe of options data
call_dollar_gammas = pd.DataFrame({'Dollar Gamma': call_dollar_gammas})
call_chain = pd.concat([call_chain, call_dollar_gammas], axis=1)

put_dollar_gammas = pd.DataFrame({'Dollar Gamma': put_dollar_gammas})
put_chain = pd.concat([put_chain, put_dollar_gammas], axis=1)

call_chain.to_excel('calls_dol_gammas.xlsx', sheet_name= ticker_text + " " + expiration_date)
put_chain.to_excel('puts_dol_gammas.xlsx', sheet_name= ticker_text + " " + expiration_date)'''

