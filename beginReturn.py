import yfinance as yf
from openpyxl import load_workbook
import datetime as dt
from datetime import date
from datetime import timedelta

#Calculates return for the next day given hedging pressure of the previous day.
def get_sheetnames_xlsx(filepath): 
    wb = load_workbook(filepath, read_only=True, keep_links=False)
    return wb.sheetnames

sheet_name = get_sheetnames_xlsx('C:/Users/Mercy/Documents/GitHub/SmartMoneyAlgo/calls_output.xlsx')
sheet_name = sheet_name[0].split()

expiration_date = sheet_name[1]

ticker_text = sheet_name[0]
ticker = yf.Ticker(ticker_text)

stock_data = ticker.history()

stock_closes = stock_data.iloc[:, 3].values.tolist()
stock_opens = stock_data.iloc[:, 0].values.tolist()
open_p = stock_opens[-1]
current_p = stock_closes[-1]
current_return =  ((current_p - open_p) / open_p) *100


wb = load_workbook("C:/Users/Mercy/Documents/GitHub/SmartMoneyAlgo/hpReturnData.xlsx")
sheet = wb.active 
last_row = len(sheet['A'])
last_row_date = sheet["A" + str(last_row)]

sheet["E" + str(last_row)] = str(round(current_return, 5)) + "%"

wb.save("C:/Users/Mercy/Documents/GitHub/SmartMoneyAlgo/hpReturnData.xlsx")
print("done")